import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.title("Calcolo Prezzo di Vendita a Litro e per Formato")

# Input margini e commissioni
commissione = st.number_input("Commissione marketplace (%):", min_value=0.0, max_value=100.0, value=14.0, step=0.1)/100
iva = st.number_input("IVA (%):", min_value=0.0, max_value=100.0, value=22.0, step=0.1)/100

margine_tipo = st.selectbox("Tipo di margine desiderato:", ["Euro", "Percentuale (%)"])
if margine_tipo == "Euro":
    margine_input = st.number_input("Margine netto desiderato (€ a litro):", min_value=0.0, value=2.0, step=0.01)
else:
    margine_input = st.number_input("Margine netto desiderato (% sul prezzo imponibile):", min_value=0.0, max_value=100.0, value=15.0, step=0.1)/100

# Upload file
uploaded_file = st.file_uploader("Carica il file Excel o CSV", type=["xlsx","csv"])
if uploaded_file:
    if uploaded_file.name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)

    df.columns = df.columns.str.strip()

    required_cols = ["Formato (L)", "Prezzo netto", "Costo spedizione"]
    if not all(col in df.columns for col in required_cols):
        st.error(f"Il file deve contenere le colonne: {required_cols}")
    else:
        st.success("Tutte le colonne richieste presenti!")

        def calcola_spedizione(spedizione, formato):
            try:
                if str(spedizione).strip().lower() == "gratis":
                    return 0.0, 0.0
                else:
                    sped_per_litro = float(spedizione) / float(formato)
                    return sped_per_litro, float(spedizione)
            except:
                return 0.0, 0.0

        df["Spedizione per litro"], df["Spedizione per formato"] = zip(*df.apply(
            lambda row: calcola_spedizione(row["Costo spedizione"], row["Formato (L)"]), axis=1))

        df["Prezzo totale a litro"] = df["Prezzo netto"] + df["Spedizione per litro"]

        # Calcolo prezzo di vendita a litro lordo
        def calcola_prezzo_lordo(prezzo_totale_litro):
            if margine_tipo == "Euro":
                denominatore = (1 / (1 + iva)) - commissione
                if denominatore <= 0:
                    return None
                return (margine_input + prezzo_totale_litro) / denominatore
            else:
                return (prezzo_totale_litro / (1 - commissione - margine_input)) * (1 + iva)

        df["Prezzo vendita a litro (€)"] = df["Prezzo totale a litro"].apply(calcola_prezzo_lordo)
        df["Prezzo vendita per formato (€)"] = df["Prezzo vendita a litro (€)"] * df["Formato (L)"]

        # Margine NETTO reale (tolto TUTTO: IVA, commissioni, costi)
        df["IVA da versare (€)"] = df["Prezzo vendita per formato (€)"] - df["Prezzo vendita per formato (€)"] / (1 + iva)
        df["Commissione marketplace (€)"] = df["Prezzo vendita per formato (€)"] * commissione

        df["Margine netto per formato (€)"] = (
            df["Prezzo vendita per formato (€)"]
            - df["IVA da versare (€)"]
            - df["Commissione marketplace (€)"]
            - (df["Prezzo totale a litro"] * df["Formato (L)"])
        )
        df["Margine netto a litro (€)"] = df["Margine netto per formato (€)"] / df["Formato (L)"]

        # Rimuovere colonne superflue e margine lordo
        colonne_da_rimuovere = [
            "Categoria", "Sottocategoria", "Nome olio", "ACEA", "Viscosità", "Tipologia",
            "Marca", "Descrizione", "Peso (Kg)", "Costo spedizione", "Mpn", "Marca veicoli",
            "Utilizzo", "Img 1", "Img 2", "Img 3", "Img 4", "Img 5", "Img 6", "Img 7",
            "competitor €/litro (spedizione inclusa)",
            "Margine reale a litro (€)", "Margine per formato (€)"  # tolte colonne margine lordo
        ]
        df = df.drop(columns=[col for col in colonne_da_rimuovere if col in df.columns])

        # Rinomina colonne
        df = df.rename(columns={
            "Prezzo netto": "Costo netto (€)",
            "Spedizione per litro": "Spedizione €/L",
            "Spedizione per formato": "Spedizione per formato (€)",
            "Prezzo totale a litro": "Costo totale €/L",
            "Prezzo vendita a litro (€)": "Prezzo vendita €/L",
        })

        st.subheader("Anteprima del file calcolato")
        st.dataframe(df.head())

        # Salvataggio Excel
        output = BytesIO()
        df.to_excel(output, index=False, sheet_name="Prezzi")
        output.seek(0)

        wb = load_workbook(output)
        ws = wb.active

        rosso_chiaro = PatternFill(start_color="FFAAAA", end_color="FFAAAA", fill_type="solid")
        verde_chiaro = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        giallo_chiaro = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

        col_costi = [
            "Costo netto (€)", "Spedizione €/L", "Spedizione per formato (€)",
            "Costo totale €/L", "IVA da versare (€)", "Commissione marketplace (€)"
        ]
        col_margini_netto = [
            "Margine netto a litro (€)", "Margine netto per formato (€)"
        ]
        col_prezzi = [
            "Prezzo vendita €/L", "Prezzo vendita per formato (€)"
        ]

        col_lettere = {cell.value: cell.column_letter for cell in ws[1]}

        # Colora colonne
        for col in col_costi:
            if col in col_lettere:
                for row in range(2, ws.max_row + 1):
                    ws[f"{col_lettere[col]}{row}"].fill = rosso_chiaro

        for col in col_margini_netto:
            if col in col_lettere:
                for row in range(2, ws.max_row + 1):
                    ws[f"{col_lettere[col]}{row}"].fill = verde_chiaro

        for col in col_prezzi:
            if col in col_lettere:
                for row in range(2, ws.max_row + 1):
                    ws[f"{col_lettere[col]}{row}"].fill = giallo_chiaro

        output_col = BytesIO()
        wb.save(output_col)
        output_col.seek(0)

        ora_corrente = datetime.now()
        nome_file = f"prezzi_vendita_{ora_corrente.strftime('%d.%m.%Y_%H.%M')}.xlsx"

        st.download_button(
            label="Scarica file Excel colorato",
            data=output_col,
            file_name=nome_file,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )